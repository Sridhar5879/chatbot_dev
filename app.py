import pandas as pd
import nltk
nltk.data.path.append('/app/nltk_data') 
import warnings
warnings.filterwarnings("ignore")
from fuzzywuzzy import fuzz
from nltk.corpus import stopwords
from nltk.corpus import wordnet
from nltk.tokenize import word_tokenize
from nltk.stem.porter import *
import stemming
from textblob import TextBlob, Word
from stemming.porter2 import stem
stop_words=set(stopwords.words('english'))

# Method to convert a series into a String
def convert(s):
  # initialization of string to ""
  new = ""
  # traverse in the string
  for x in s:
    new += x+' '
    # return string
  return new

# Method for writing into an existing excelsheet
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=2,truncate_sheet=True,**to_excel_kwargs):

    from openpyxl import load_workbook
    import pandas as pd
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
#Reading the document and the relevant fields
df=pd.read_excel("chatbot - CGL.xlsx")
SD = pd.read_excel("CGL Retailer Details.xlsx")
dfq=df['Question']
dfa=df['Answer']

#global variables
userid_global = ""

#Lemmatization of "Questions"
from textblob import Word
dfq1 = dfq.apply(lambda x: " ".join([Word(word).lemmatize() for word in x.split()]))
#Removal of stop words
from nltk.corpus import stopwords
stop = stopwords.words('english')
dfq1 = dfq1.apply(lambda x: " ".join(x for x in x.split() if x not in stop))
  
# Changing the "Questions" to lower case
dfq_1 = [w.lower() for w in dfq1]
# "Answers" have been converted to a list
dfa1=[w for w in dfa]

# Basic code required to run the app on heroku
from flask import Flask, render_template, request, json, jsonify, make_response
import requests
import pdb
app = Flask(__name__)


@app.route('/', methods=['GET','POST'])
def order_status():
  if request.method == 'POST':
      #Taking the input query from user and converting it to an usable string
      try:
          userid = int(request.form.get('ui_query'))
          if userid in list(SD['ID']):
              Retailer = SD[SD['ID'] == userid]
          else:
              aa = {}
              aa['input'] = int(request.form.get('ui_query'))
              aa['result'] = "Sorry but your User Id did not match with any of our records, please try again"
              print('Sorry but your User Id did not match with any of our records, please try again')
              return aa
          
          aa = {}
          aa['input'] = int(request.form.get('ui_query'))
          aa['sss'] = userid
          append_df_to_excel('Book2.xlsx', Retailer, header=1, index=False, startrow=0, sheet_name='Data')
          names = list(Retailer['Name'])
          name =names[0]
          if userid == 932245:
              result = "Hi %s, India Sales '\n' please select a report from the following <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report GH.xlsx?raw=true'>IndiaSalesReport </a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report NI.xlsx?raw=true'>NorthIndiaSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report SI.xlsx?raw=true'>SouthIndiaSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report PJB.xlsx?raw=true'>PunjabSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report MH.xlsx?raw=true'>MaharastraSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report TG.xlsx?raw=true'>TelanganaSalesReport</a>  "%name
          elif userid == 941795:
              result = "Hi %s, South India Sales '\n' please select a report from the following <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report SI.xlsx?raw=true'>SouthIndiaSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report TG.xlsx?raw=true'>TelanganaSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report MH.xlsx?raw=true'>MaharastraSalesReport</a> "%name
          elif userid == 934771:
              result = "Hi %s, North India Sales '\n'please select a report from the following <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report NI.xlsx?raw=true'>NorthIndiaSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report PJB.xlsx?raw=true'>PunjabSalesReport</a> "%name
          elif userid == 415250:
              result = "Hi %s, Maharastra Sales '\n'please select a report from the following <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report MH.xlsx?raw=true'>MaharastraSalesReport</a> <a href = 'https://github.com/Sridhar5879/chatbot_dev/blob/master/Sales report Mum.xlsx?raw=true'>MumbaiSalesReport</a> "%name
          aa['result'] = result
          return aa
      except ValueError:
          query = request.form.get('ui_query')
          query1 = pd.Series(query)
          query2 = query1.apply(lambda x: " ".join([Word(word).lemmatize() for word in x.split()]))
          query2=  [w for w in query2 if not w in stop]
          
          # Converting query2 into a string
          k=convert(query2)
          #Implementation fuzzywuzzy algorithm to find the closest match
          from fuzzywuzzy import process
          ## To Get Related questions based on ratio
          choices_dict = {idx: el for idx, el in enumerate(dfq_1)}
          Ratios = process.extract(k,choices_dict,limit=3)
          ChatReply=(tuple(Ratios[0]))
          j=(tuple(Ratios[1]))
          l=(tuple(Ratios[2]))
          if ChatReply[1]<70:
              aa = {}
              aa['input'] = request.form.get('ui_query')
              aa['result'] = "Sorry but your query did not match with any of our records, please try with another query"
                  
          elif ChatReply[1]>=70 and ChatReply[2] <= 48:
              aa = {}
              aa['input'] = request.form.get('ui_query')
              aa['result'] = dfa[ChatReply[2]]
                  
          else:
              aa = {}
              aa['input'] = request.form.get('ui_query')
              aa['result'] = dfa[ChatReply[2]]
              print('\n','NEAREST MATCH','\n''\n',dfq[ChatReply[2]],'\n',dfa[ChatReply[2]],'\n','\n','DO YOU MEAN?','\n',dfq[j[2]],'\n',dfq[l[2]])
          return aa
      
  else:
      return render_template('chat.html')

if __name__ == "__main__":
  app.run(host='127.0.0.1')